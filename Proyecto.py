import RPi.GPIO as GPIO #Librería para el control de los motores a pasos y el servomotor
from RpiMotorLib import RpiMotorLib #Librería para motores a pasos
from RpiMotorLib import rpiservolib #Librería para servomotor
import time #Librería de tiempo
from time import sleep #Libraría para sleep
import serial #Librería para configuración y adquisición de datos de dispositivos seriales
import openpyxl #Librería para hojas de datos
from tkinter import * #Librería para la interfaz
from tkinter import ttk #Librería para la interfaz

################## Captura de datos TESA ##################

serTESA=serial.Serial('/dev/ttyUSB1', baudrate=1200, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN,
                      stopbits=serial.STOPBITS_TWO, xonxoff=True, timeout=0.5) #Configuración de puerto

def DatosTESA(serTESA): 
    detenerse=0 #Constante para while que captura dato
    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(30) #Lectura de 30 bytes
            if data == '':
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serTESA) #Llamada de la función
        if data != b'': #Comparación de datos recibidos, vacío hasta que se de la medición
            medicion=float(data) #Pasando de string a float
            MedicionBloque=medicion #Guardando dato en lista
            detenerse = 1 #Condición para salir del while
    return MedicionBloque

################## Captura de datos Fluke ##################

serFluke=serial.Serial('/dev/ttyUSB2', baudrate=9600, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE,
                       stopbits=serial.STOPBITS_ONE, xonxoff=True, timeout=0.5) #Configuración de puerto

def DatosFluke(serFluke):
    serFluke.write(b'READ?1\r\n') #Envío de instrucción para capturar dato de temperatura 1
    serFluke.write(b'READ?2\r\n') #Envío de instrucción para capturar dato de temperatura 2
    serFluke.write(b'READ?3\r\n') #Envío de instrucción para capturar dato de temperatura 3
    serFluke.write(b'READ?4\r\n') #Envío de instrucción para capturar dato de temperatura 4

    detenerse=0 #Constante para while que captura dato
    
    MedicionTemp1=0 #Creación de variable para almacenar mediciones de temperatura 1
    MedicionTemp2=0 #Creación de variable para almacenar mediciones de temperatura 2
    MedicionTemp3=0 #Creación de variable para almacenar mediciones de temperatura 3
    MedicionTemp4=0 #Creación de variable para almacenar mediciones de temperatura 4


    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(32) #Lectura de 32 bytes
            if data == '':
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serFluke) #Llamada de la función   
        if data != b'': #Comparación de datos recibidos, vacío hasta que se de la medición
            todas=data.split()#Separar los 4 datos en una lista
            MedicionTemp1=float(todas[0]) #Guardando temperatura 1 en lista
            MedicionTemp2=float(todas[1]) #Guardando temperatura 2 en lista
            MedicionTemp3=float(todas[2]) #Guardando temperatura 3 en lista
            MedicionTemp4=float(todas[3]) #Guardando temperatura 4 en lista
            detenerse = 1  #Condición para salir del while
    return MedicionTemp1, MedicionTemp2, MedicionTemp3, MedicionTemp4

################## Captura de datos Vaisala ##################

serVaisala=serial.Serial('/dev/ttyUSB0', baudrate=4800, bytesize=serial.SEVENBITS,
                         parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5) #Configuración de puerto
#serVaisala=serial.Serial('/dev/ttyAMA0', baudrate=4800, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5)
#serVaisala=serial.Serial('/dev/ttyUSB0', baudrate=19200, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout= 0.5)

#serVaisala.write(b'RUN\r\n')
#serVaisala.write(b'R\r/\n')
#serVaisala.write(b'form "P=" 4.2 P " " U6 \t "T=" t " " U3 \t "RH=" 4.2 rh " " U5 \r\n')
#serVaisala.write(b'form 4.2 P " " \t 4.2 t " " \t 4.2 rh " " \r\n')

def DatosVaisala(serVaisala):
    serVaisala.write(b'SEND\r\n') #Envío de instrucción para capturar datos del Vaisala
    
    detenerse=0 #Constante para while que captura dato
    
    DatoPresVaisala=0 #Creación de variable para almacenar mediciones de presión atmosférica
    DatoTempVaisala=0 #Creación de variable para almacenar mediciones de temperatura ambiente
    DatoHumeVaisala=0 #Creación de variable para almacenar mediciones de humedad relativa
    
    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(30)
            if data == '':
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serVaisala) #Llamada de la función
        if data != b'': #Comparación de datos recibidos, vacío hasta que se de la medición
            todos=data.split()#Separar los 4 datos en una lista
            DatoPresVaisala=float(todos[1]) #Guardando presión atmosférica en lista
            DatoTempVaisala=float(todos[2]) #Guardando temperatura en lista
            DatoHumeVaisala=float(todos[3]) #Guardando humedad relativa en lista
            detenerse = 1 #Condición para salir del while
    return DatoPresVaisala, DatoTempVaisala, DatoHumeVaisala

################## Activar Pedal ##################

servo_pin = 26 #Pin que envía la señal al servomotor

def ActivaPedal(servo_pin): 

    myservotest = rpiservolib.SG90servo("servoone", 50, 2, 12) #Parámetros del servomotor

    myservotest.servo_move(servo_pin, 2.3, .5, False, .01) #Movimiento a posición 2.3
    myservotest.servo_move(servo_pin, 7.5, .5, False, .01) #Movimiento a posición 7.5


################## Secuencia completa - Plantilla1 ##################

def Completa1(repeticiones):

    listaMediciones=[]

    for i in range(repeticiones):

        ActivaPedal(servo_pin) #Sube palpador
                        
        mymotortest1.motor_go(False, "Half", 208, .003, False, 2) #Mov de HOME a 1

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del patrón en posición 1 (centro)

        mymotortest1.motor_go(True, "Half", 416, .003, False, 2) #Mov de 1 a 2

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 2 (esquina)

        mymotortest1.motor_go(False, "Half", 96, .003, False, 2) #Mov1 de 2 a 3
        mymotortest2.motor_go(False, "Full", 400, .005, False, 1) #Mov2 de 2 a 3
        mymotortest1.motor_go(True, "Half", 180, .003, False, 1) #Mov3 de 2 a 3

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del calibrando en posición 3 (esquina)
            
        mymotortest1.motor_go(False, "Half", 181, .003, False, 2) #Mov de 3 a 4

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 4 (esquina)

        mymotortest2.motor_go(True, "Full", 800, .005, False, 2) #Mov de 4 a 5

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 5 (esquina)

        mymotortest1.motor_go(True, "Half", 178, .003, False, 2) #Mov de 5 a 6

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 6 (esquina)

        mymotortest1.motor_go(False, "Half", 178, .003, False, 2) #Mov de 6 a 5
        mymotortest2.motor_go(False, "Full", 400, .005, False, 1) #Mov de 5 a Esp2
        mymotortest1.motor_go(False, "Half", 332, .003, False, 1) #Mov de Esp2 a 1

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)

        mymotortest1.motor_go(True, "Half", 208, .003, False, 2) #Mov de 1 a HOME

        ActivaPedal(servo_pin) #Baja palpador

    return listaMediciones

################## Secuencia completa - Plantilla2 ##################

def Completa2(repeticiones):

    for i in range(repeticiones):

        ActivaPedal(servo_pin) #Sube palpador

        mymotortest1.motor_go(False, "Half", 208, .003, False, 2) #Mov de HOME a 1

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)

        mymotortest1.motor_go(True, "Half", 416, .003, False, 2) #Mov de 1 a 2

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 2 (esquina)
        
        mymotortest1.motor_go(False, "Half", 96, .003, False, 2) #Mov1 de 2 a 3
        mymotortest2.motor_go(False, "Full", 337, .005, False, 1) #Mov2 de 2 a 3
        mymotortest1.motor_go(True, "Half", 180, .003, False, 1) #Mov3 de 2 a 3

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 3 (esquina)
        
        mymotortest1.motor_go(False, "Half", 181, .003, False, 2) #Mov de 3 a 4

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 4 (esquina)

        mymotortest2.motor_go(True, "Full", 683, .005, False, 2) #Mov de 4 a 5

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 5 (esquina)

        mymotortest1.motor_go(True, "Half", 178, .003, False, 2) #Mov de 5 a 6

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 6 (esquina)

        mymotortest1.motor_go(False, "Half", 178, .003, False, 2) #Mov de 6 a 5
        mymotortest2.motor_go(False, "Full", 342, .005, False, 1) #Mov de 5 a Esp2
        mymotortest1.motor_go(False, "Half", 332, .003, False, 1) #Mov de Esp2 a 1

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)

        mymotortest1.motor_go(True, "Half", 208, .003, False, 2) #Mov de 1 a HOME
        
        ActivaPedal(servo_pin) #Baja palpador

    return listaMediciones

################## Secuencia centros ##################

def Centros(repeticiones):

    listaMediciones=[]

    for i in range(repeticiones):

        ActivaPedal(servo_pin) #Sube palpador

        mymotortest1.motor_go(False, "Half", 203, .003, False, 2) #Mov de HOME a 1

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)

        mymotortest1.motor_go(True, "Half", 406, .003, False, 2) #Movimiento de punto1 a punto2

        ActivaPedal(servo_pin) #Baja palpador
        sleep(tiempoEstabilizacion) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA(serTESA) #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 2 (esquina)

        mymotortest1.motor_go(False, "Half", 203, .003, False, 2) #Movimiento de punto2 a HOME

    mymotortest1.motor_go(False, "Half", 203, .003, False, 2) #Movimiento de HOME a punto1

    ActivaPedal(servo_pin) #Baja palpador
    sleep(tiempoEstabilizacion) #Tiempo de estabilización
    ActivaPedal(servo_pin) #Sube palpador
    MedicionBloque=DatosTESA(serTESA) #Llama función TESA
    listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)

    mymotortest1.motor_go(True, "Half", 203, .003, False, 2) #Mov de 1 a HOME

    ActivaPedal(servo_pin) #Baja palpador

    return listaMediciones

################## Creación de hoja de datos - Secuencia completa ##################

def HojaDatosCom(repeticiones):
    
    wb = openpyxl.Workbook() #Creación de libro
    hoja = wb.active #Activar hoja
    hoja.title = "Datos de calibración" #Definir nombre de primer hoja
    hoja1 = wb.create_sheet("Información de servicio") #Definir nombre de segunda hoja

    encabezado1=[] #Creación de lista vacía para guardar encabezado 1
    encabezado1=['','','','','']
    for i in range(repeticiones): #Se harán espacios para cada medición
        encabezado1.append('Medición #'+str(i+1))
        for j in range(6):
            encabezado1.append('')

    encabezado1.append('Medición Final') 

    hoja.append(encabezado1) #Crea la fila del encabezado 1 con los títulos

    encabezado2=[] #Creación de lista vacía para guardar encabezado 2
    encabezado2=['Número de bloque','Fecha','Hora','Identificación del bloque', 'Valor nominal del bloque (mm o pulg)']

    for i in range(repeticiones): #Se harán espacios para los valores de cada medición
        encabezado2.append('Valor del Patrón en posición 1 (Centro) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 2 (Centro) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 3 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 4 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 5 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 6 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Patron en posición 1 (Centro) medición #'+str(i+1)+' (µm o µpulg')

    encabezado2.append('Valor del Patrón en posición 1 (Centro) medición #'+str(repeticiones+1)+' (µm o µpulg)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #1 (°C)') #Encabezado para temperaturas al inicio de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)') #Encabezado para temperaturas al final de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)')

    encabezado2.append('Humedad Relativa Promedio Vaisala (%)') #Encabezado para promedio de humedad relativa

    hoja.append(encabezado2) #Crea la fila del encabezado 2 con los títulos

################## Creación de hoja de datos - Secuencia centros ##################

def HojaDatosCen(repeticiones):
    
    wb = openpyxl.Workbook() #Creación de libro
    hoja = wb.active #Activar hoja
    hoja.title = "Datos de calibración" #Definir nombre de primer hoja
    hoja1 = wb.create_sheet("Información de servicio") #Definir nombre de segunda hoja

    encabezado1=[] #Creación de lista vacía para guardar encabezado 1
    encabezado1=['','','','','']
    for j in range(repeticiones): #Se harán espacios para cada medición
        encabezado1.append('Medición #'+str(i+1))
        encabezado1.append('')

    encabezado1.append('Medición Final') 

    hoja.append(encabezado1) #Crea la fila del encabezado 1 con los títulos

    encabezado2=[] #Creación de lista vacía para guardar encabezado 2
    encabezado2=['Número de bloque','Fecha','Hora','Identificación del bloque', 'Valor nominal del bloque (mm o pulg)']

    for j in range(repeticiones): #Se harán espacios para los valores de cada medición
        encabezado2.append('Valor del Patrón medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando medición #'+str(i+1)+' (µm o µpulg)')

    encabezado2.append('Valor del Patrón medición #'+str(repeticiones+1)+' (µm o µpulg)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #1 (°C)') #Encabezado para temperaturas al inicio de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)') #Encabezado para temperaturas al final de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(repeticiones+1)+' (°C)')

    encabezado2.append('Humedad Relativa Promedio Vaisala (%)') #Encabezado para promedio de humedad relativa

    hoja.append(encabezado2) #Crea la fila del encabezado 2 con los títulos

################## Configuración de los motores a pasos ##################

GPIO_pins1 = (22, 27, 17) #pines de modo para el motor1
direction1 = 9 #pin de dirección para el motor1
step1 = 11 #pin de step para el motor1

GPIO_pins2 = (5, 6, 13) #pines de modo para el motor2
direction2 = 20 #pin de dirección para el motor2
step2 = 21 #pin de step para el motor2

EN_pin = 24 #pin de enable

mymotortest1 = RpiMotorLib.A4988Nema(direction1, step1, GPIO_pins1, "A4988") #Parámetros del motor1
mymotortest2 = RpiMotorLib.A4988Nema(direction2, step2, GPIO_pins2, "A4988") #Parámetros del motor2

GPIO.setup(EN_pin, GPIO.OUT)                                                                                                                                                
GPIO.output(EN_pin, GPIO.LOW) #Enable debe estar en LOW

################## Llenar o cargar datos de identificación y valor nominal de bloques ##################

identificacionBloques='falta'
valorNominalBloques='falta'
        
################## Selección de secuencia ##################

secuencia=int(input('Escoger secuencia'))
cantidadBloques=int(input('Cantidad de bloques'))
BloqueActual=1
repeticiones=int(input('Cantidad de repeticiones: ')) #Petición del número de repeticiones que se harán de la secuencia
tiempoEstabilizacion=int(input('Tiempo de estabilización'))

listaDatos=[]

listaDatos.extend(BloqueActual) #Número de bloque
listaDatos.extend(time.strftime("%d/%m/%y")) #Fecha
listaDatos.extend(time.strftime("%H:%M:%S")) #Hora
listaDatos.extend(identificacionBloques) #Identificación del bloque
listaDatos.extend(valorNominalBloques) #Valor Nominal del bloque

DatoPresVaisala, DatoTempVaisala, DatoHumeVaisala=DatosVaisala(serVaisala) #Llama función Vaisala

DatoPresVaisala1, DatoTempVaisala1, DatoHumeVaisala1=DatosVaisala(serVaisala) #Llama función Vaisala
PromedioHumedad=(DatoHumeVaisala+DatoHumeVaisala1)/2

################## Interfaz ##################

window=Tk()
window.title("Calibración de bloques")#Nombre de la ventana
window.resizable(0,0)#No permite cambiar dimensiones de la venta
window.geometry("1000x800") #Tamaño de la ventana

secEscogida= IntVar()
planEscogida= IntVar()

def irCalibra():
    FrameInfo.pack_forget()
    FrameCalibra.pack(side="top", fill="both", expand=True)

def obtenerSecuencia():
    Secuencia=secEscogida.get()
    Plantilla=planEscogida.get()
    return Secuencia,Plantilla

def bloqueActual(BloqueActual):
    BloqueActual=BloqueActual+1
    return BloqueActual
   
# def selectorSecuencia():
#     listaDatos=[]
#     Secuencia,Plantilla=obtenerSecuencia()
#     listaDatos.extend(BloqueActual) #Número de bloque
#     listaDatos.extend(time.strftime("%d/%m/%y")) #Fecha
#     listaDatos.extend(time.strftime("%H:%M:%S")) #Hora
#     listaDatos.extend(identificacionBloques) #Identificación del bloque
#     listaDatos.extend(valorNominalBloques) #Valor Nominal del bloque
#     if BloqueActual==1:
#         DatoPresVaisala1, DatoTempVaisala1, DatoHumeVaisala1=DatosVaisala(serVaisala) #Llama función Vaisala
#     MedicionTemp1, MedicionTemp2, MedicionTemp3, MedicionTemp4=DatosFluke(serFluke) #Llama función Fluke
#     if Secuencia==1 and Plantilla==1: #Secuencia completa - plantilla 1
#         listaMediciones=Completa1(repeticiones)
#     if Secuencia==1 and Plantilla==2: #Secuencia completa - plantilla 2
#         listaMediciones=Completa2(repeticiones)
#     if Secuencia==2: #Secuencia centros
#         listaMediciones=Centros(repeticiones)
#     BloqueActual=bloqueActual(BloqueActual)
#     MedicionTemp5, MedicionTemp6, MedicionTemp7, MedicionTemp8=DatosFluke(serFluke) #Llama función Fluke
#     if BloqueActual==cantidadBloques:
#         DatoPresVaisala2, DatoTempVaisala2, DatoHumeVaisala2=DatosVaisala(serVaisala) #Llama función Vaisala
#         PromedioHumedad=(DatoHumeVaisala+DatoHumeVaisala1)/2

def selectorSecuencia():
    Secuencia,Plantilla=obtenerSecuencia()
    if Secuencia==1 and Plantilla==1: #Entra a venta de medición de diametro
        print('Secuencia completa plantilla 1')
    if Secuencia==1 and Plantilla==2: #Entra a venta de medición de longitud
        print('Secuencia completa plantilla 2')
    if Secuencia==2: #Entra a venta de medición de longitud
        print('Secuencia centros')

FrameInfo=Frame(window) #Crea el frame donde estará la información de servicio
FrameCalibra=Frame(window) #Crea el frame donde se realizará el proceso de calibración

################## Frame Información de servicio ##################

FrameInfo.pack(side="top", fill="both", expand=True) #Empieza el frame de información

Label(FrameInfo, text="Información servicio").place(x=10, y=10)

Label(FrameInfo, text="Número de certificado").place(x=10, y=40)
numeroCertificado=Entry(FrameInfo).place(x=200, y=40)

ttk.Button(FrameInfo, text='Siguiente', command=irCalibra).place(x=855, y=740) #Boton para ir a frame de calibración

################## Frame Calibración ##################

Label(FrameCalibra, text="Secuencia: ").place(x=10, y=10)
Radiobutton(FrameCalibra, text="Completa", variable=secEscogida,value=1, command=obtenerSecuencia).place(x=10, y=40)
Radiobutton(FrameCalibra, text="Centros", variable=secEscogida,value=2, command=obtenerSecuencia).place(x=10, y=70)

Label(FrameCalibra, text="Plantilla: ").place(x=10, y=130)
Radiobutton(FrameCalibra, text="1", variable=planEscogida,value=1, command=obtenerSecuencia).place(x=10, y=160)
Radiobutton(FrameCalibra, text="2", variable=planEscogida,value=2, command=obtenerSecuencia).place(x=10, y=190)

Label(FrameCalibra, text="Tiempo de estabilización: ").place(x=10, y=250)
tiempoEstabilizacion=Entry(FrameCalibra).place(x=200, y=250)

Label(FrameCalibra, text="Repeticiones: ").place(x=10, y=280)
repeticiones=Entry(FrameCalibra).place(x=200, y=280)

boton2=ttk.Button(FrameCalibra, text='Iniciar calibración', command=selectorSecuencia).place(x=10, y=500) #Boton para iniciar calibración

boton3=ttk.Button(FrameCalibra, text='Pausar calibración', command=selectorSecuencia).place(x=445, y=500) #Boton para parar secuencia #guardarParametros

boton4=ttk.Button(FrameCalibra, text='Reanudar calibración', command=selectorSecuencia).place(x=840, y=500) #Boton para parar secuen

ttk.Button(window, text='Salir', command=quit).pack(side=BOTTOM) #Boton de salir

window.mainloop()

GPIO.cleanup()

        
         

