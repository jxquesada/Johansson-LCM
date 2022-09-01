from tkinter import *
from tkinter import ttk
import xlrd
from tkinter import filedialog
import time
from datetime import date
import RPi.GPIO as GPIO             #Biblioteca para el control de los motores a pasos y el servomotor
from RpiMotorLib import RpiMotorLib #Biblioteca para motores a pasos
from RpiMotorLib import rpiservolib #Biblioteca para servomotor
from time import sleep              #Biblioteca para sleep
import serial                       #Biblioteca para configuración y adquisición de datos de dispositivos seriales
import openpyxl                     #Biblioteca para hojas de datos
from openpyxl import load_workbook  #Biblioteca para cargar excel ya existente
import smtplib, ssl

################## Creación de ventana ##################

window=Tk()
window.title("Calibración de bloques")          #Nombre de la ventana
window.resizable(0,0)                           #No permite cambiar dimensiones de la venta
                                   
window.geometry("1600x900")                     #Tamaño de la ventana

################## Setear posición del disco giratorio ##################
def gohome():
    #sleep(1)
    GPIO.output(EN_pin, GPIO.LOW)       #Enable debe estar en LOW para accionar motores
    GPIO.output(sleepMot3, GPIO.HIGH)       #Sleep debe estar en HIGH para funcionar
    sensorOut=GPIO.input(slotted_pin)
    print("Estado del sensor infrarrojo", sensorOut)
    sleep(5)
    steps=0                         #Define el contador de steps en 0
    print('Dirigiéndose a home')
    
    while sensorOut ==0:
        #print("Moviéndose y monitoreando el sensor infrarrojo")
        sensorOut=GPIO.input(slotted_pin)
                                    #Cuando no esté en home avance
        motor3.motor_go(True,       #Clockwise
                        "Full",     #Tipo de step
                        1,          #Cantidad de steps
                        .005,         #Delay del step (s)
                        False,      #No imprimir salida
                        .05)        #Delay inicial (s)
        steps+=1                    #Suma el paso que dio
                                    #Corregimos la posición proporcionalmente al recorrido
                                    #esto debido a la inercia encontrada
    print("Cantidad de pasos hasta home: ", steps)
    devolverse=0.025*steps+15        #Estimar cuánto devolverse
    devolverse=round(devolverse, 0) #Redondea a número entero
    devolverse=int(devolverse)
    print("Cantidad de pasos de ajuste a home: ", devolverse)
    sleep(5)
    for count in range(devolverse):
        motor3.motor_go(False,       #Clockwise
                        "Full",     #Tipo de step
                        1,          #Cantidad de steps
                        .005,         #Delay del step (s)
                        False,      #No imprimir salida
                        .05)        #Delay inicial (s)
    #global posicionStep
    posicionStep=0                  #Defino la posición home como angulo 0
    GPIO.output(sleepMot3, GPIO.LOW)       #Sleep debe estar en LOW para deshabilitarse
    print("Estoy en casa")
    GPIO.output(EN_pin, GPIO.HIGH)       #Inhabilita los motores
    return posicionStep
################## Configuración de entradas/salidas ##################

GPIO_pins1 = (22, 27, 17)           #pines de modo para el motor1
direction1 = 9                      #pin de dirección para el motor1
step1 = 11                          #pin de step para el motor1

GPIO_pins2 = (5, 6, 13)             #pines de modo para el motor2
direction2 = 20                     #pin de dirección para el motor2
step2 = 21                          #pin de step para el motor2

EN_pin = 24                         #pin de enable

GPIO_pins3 = (14, 15, 18)           #Pines de modo de paso
direction3 = 19                     #Pin de sentido de giro
step3 = 16                          #Pin de dar paso
sleepMot3=12                        #Pin para controlar el sleep del motor de ordenamiento
                                    #Si está en 1 está activo, en 0 está en sleep

motor3 = RpiMotorLib.A4988Nema(direction3, step3, GPIO_pins3, "A4988") #Parámetros del motor

mymotortest1 = RpiMotorLib.A4988Nema(direction1, step1, GPIO_pins1, "A4988") #Parámetros del motor1
mymotortest2 = RpiMotorLib.A4988Nema(direction2, step2, GPIO_pins2, "A4988") #Parámetros del motor2

GPIO.setup(EN_pin, GPIO.OUT)                                                                                                                                                
GPIO.output(EN_pin, GPIO.HIGH)       #Modo seguro, motores inhabilitados

GPIO.setup(sleepMot3, GPIO.OUT)                                                                                                                                                
GPIO.output(sleepMot3, GPIO.LOW)       #Sleep debe estar en LOW para deshabilitarse

slotted_pin = 4                     #Pin para el sensor infrarrojo
GPIO.setmode(GPIO.BCM)              #Numeración Broadcom
GPIO.setup(slotted_pin, GPIO.IN)    #Se define como entrada el sensor

posicionStep=0                      #Variable de posición angular del disco
required=0                          #Variable de pasos requeridos par llegar
                                    #a la posicion deseada
listo=0                             #Variable que determina cuando terminó

#gohome()                            #gire el disco hasta home porque se inició el programa

################## Creación de hoja de datos - Información de servicio ##################

################## Creación de hoja de datos - Secuencia completa ##################

def HojaDatosCom(Repeticiones):
    
    global wb

    hoja = wb.active                            #Activar hoja

    encabezado1=[]                              #Creación de lista vacía para guardar encabezado 1
    encabezado1=['','','','','']
    for i in range(int(repeticiones.get())):    #Se harán espacios para cada medición
        encabezado1.append('Medición #'+str(i+1))
        for j in range(6):
            encabezado1.append('')

    encabezado1.append('Medición Final') 

    hoja.append(encabezado1)                    #Crea la fila del encabezado 1 con los títulos

    encabezado2=[]                              #Creación de lista vacía para guardar encabezado 2
    encabezado2=['Número de bloque','Fecha','Hora','Identificación del bloque', 'Valor nominal del bloque (mm o pulg)']

    for i in range(int(repeticiones.get())):    #Se harán espacios para los valores de cada medición
        encabezado2.append('Valor del Patrón en posición 1 (Centro) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 2 (Centro) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 3 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 4 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 5 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando en posición 6 (esquina) medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Patron en posición 1 (Centro) medición #'+str(i+1)+' (µm o µpulg')

    encabezado2.append('Valor del Patrón en posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (µm o µpulg)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #1 (°C)')
                                                #Encabezado para temperaturas al inicio de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')
                                                #Encabezado para temperaturas al final de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')

    encabezado2.append('Humedad Relativa Promedio Vaisala (%)') #Encabezado para promedio de humedad relativa

    hoja.append(encabezado2) #Crea la fila del encabezado 2 con los títulos
    
    wb.save(numeroCertificado.get()+'.xlsx')

################## Creación de hoja de datos - Secuencia centros ##################

def HojaDatosCen(Repeticiones):
    
    global wb
    
    hoja = wb.active                                #Activar hoja
    
    encabezado1=[]                                  #Creación de lista vacía para guardar encabezado 1
    encabezado1=['','','','','']
    for i in range(int(repeticiones.get())):        #Se harán espacios para cada medición
        encabezado1.append('Medición #'+str(i+1))
        encabezado1.append('')

    encabezado1.append('Medición Final') 

    hoja.append(encabezado1)                        #Crea la fila del encabezado 1 con los títulos

    encabezado2=[]                                  #Creación de lista vacía para guardar encabezado 2
    encabezado2=['Número de bloque','Fecha','Hora','Identificación del bloque', 'Valor nominal del bloque (mm o pulg)']

    for i in range(int(repeticiones.get())):        #Se harán espacios para los valores de cada medición
        encabezado2.append('Valor del Patrón medición #'+str(i+1)+' (µm o µpulg)')
        encabezado2.append('Valor del Calibrando medición #'+str(i+1)+' (µm o µpulg)')

    encabezado2.append('Valor del Patrón medición #'+str(int(repeticiones.get())+1)+' (µm o µpulg)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #1 (°C)')
                                                    #Encabezado para temperaturas al inicio de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #1 (°C)')

    encabezado2.append('Temperatura del Patrón durante la toma del Valor del Patrón en  posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')
                                                    #Encabezado para temperaturas al final de la secuencia
    encabezado2.append('Temperatura del Calibrando durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')
    encabezado2.append('Temperatura ambiente dentro de la cámara durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')
    encabezado2.append('Temperatura del Calibrador de bloques durante la toma del Valor del Patrón en posición 1 (Centro) medición #'+str(int(repeticiones.get())+1)+' (°C)')

    encabezado2.append('Humedad Relativa Promedio Vaisala (%)')
                                                    #Encabezado para promedio de humedad relativa

    hoja.append(encabezado2)                        #Crea la fila del encabezado 2 con los títulos
    
    wb.save(numeroCertificado.get()+'.xlsx')
    
###################### Movimientos del órdenes de trabajo#########################################
def obtenerAnguloBloque(valorNominalBloque):         #funcion para extraer el angulo al que debe moverse el disco
    global desired
    global posicionStep #NUEVO
    print("Moviendo disco hacia siguiente pareja de bloques a calibrar")
    print("Siguiente bloque: ", valorNominalBloque)
    #loc = (r"C:Descargas\Ubicaciones.xlsx")
    #loc = (r"Ubicaciones.xlsx")
    loc = ("/home/pi/Ubicaciones.xlsx")                          #path al archivo de posiciones

                                                            #path al archivo de posiciones
    wb_obj = openpyxl.load_workbook(loc)                    #NUEVO
 
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active

     
    # Note: The first row or
    # column integer is 1, not 0.
     

        #ubicaciones = xlrd.open_workbook(loc)
    #sheetUbicaciones = ubicaciones.sheet_by_index(0)
    #sheetUbicaciones.cell_value(0, 0)
    for i in range(2, (sheet_obj.max_row+1)):                 #recorre las ubicaciones de ordenamiento hasta encontrar el valor nominal  NUEVO
        cell_obj = sheet_obj.cell(row = i, column = 1)
        #print("Fila revisada: ", cell_obj.value)
        print(type(cell_obj.value))
        if float(cell_obj.value)==float(valorNominalBloque):   #nuevo
            print("Encontré el bloque solicitado")
            cell_objEncontrado = sheet_obj.cell(row = i, column = 2)
            desired=cell_objEncontrado.value                   #extrae la posicion en steps requerida #Nuevo
            desired=desired/1.8
            print("Moviéndose para posicionar el bloque solicitado hasta la posición de (steps):")
            gire(posicionStep, desired)                     #Muévase a esa posición angular
            
            
            



def gire(posicionSteps, desired): #NUEVO
        global listo
        global posicionStep #NUEVO
        #print("La posición actual del disco (steps) es: ", posicionSteps)
        #print("La posición requerida del disco (steps) es: ", desired)
        listo=0
        
        GPIO.output(sleepMot3, GPIO.HIGH)       #Sleep debe estar en HIGH para habilitarse
        GPIO.output(EN_pin, GPIO.LOW)       #habilita los motores
        
        if desired > posicionSteps:  #Si debe moverse en sentido horario
            required=desired-posicionSteps
            print("Steps requeridos para ir a nuevo bloque",required)
            compensar=required*0.025#Compensar pasos por inercia del disco
            compensar=round(compensar,0)
                                    #La cantidad de pasos debe ser un número entero
            avanzar=required-compensar
            
            for a in range(avanzar):
                motor3.motor_go(True,   #Clockwise
                            "Full",     #Tipo de step
                            1,          #Cantidad de steps
                            .005,         #Delay del step (s)
                            False,      #No imprimir salida
                            .05)        #Delay inicial (s)
                            
            #global posicionStep
            posicionStep=posicionSteps+required
                                    #se actualiza la posicion
            #global listo
            listo=1                 #Finalizó de moverse
        elif desired < posicionSteps:#si debe moverse en sentido antihorario
            required=posicionSteps-desired
            print("Steps requeridos para ir a nuevo bloque",required)
            compensar=required*0.025#Compensar pasos por inercia del disco
            compensar=round(compensar,0)
                                    #La cantidad de pasos debe ser un número entero
            avanzar=required-compensar
            
            for a in range(avanzar):                       
                motor3.motor_go(False,  #Counter Clockwise
                            "Full",     #Tipo de step
                            1,          #Cantidad de steps
                            .005,         #Delay del step (s)
                            False,      #No imprimir salida
                            .05)        #Delay inicial (s)
                            
            #global posicionStep
            posicionStep=posicionSteps-required
                                    #se actualiza la posicion

            #global listo
            listo=1                 #Finalizó de moverse
        elif desired == posicionSteps:
            #global listo
            listo=1                 #Ya estaba en la posicion deseada
            posicionStep=posicionSteps
        else:
            #global listo
            listo=1                 #En caso de que no atienda ninguno
            posicionStep=posicionSteps
                                    #de los casos anteriores no se mueva
            
        GPIO.output(sleepMot3, GPIO.LOW)       #Sleep debe estar en LOW para deshabilitarse
        GPIO.output(EN_pin, GPIO.HIGH)       #Inhabilita los motores
        
        print("Nueva posicion del disco",posicionStep)
        return posicionStep, listo

def openfilename():                 #Abre el buscador de archivos
    filename = filedialog.askopenfilename(title ='Cargar archivo') #Busca la imagen en carpeta
    return filename

def angle2steps(angle):
    steps=int(angle/1.8)                #Convierte la posición deseada en pasos del stepper
    return steps

################## Creación de variables ##################

secEscogida= IntVar()
planEscogida= IntVar()
escogerFolio= IntVar()
escogerUnidades= IntVar()
escogerMaterial= IntVar()
escogerMaterialPatron= IntVar()
escogerGrado= IntVar()
escogerUsuario= IntVar()
escogerRevisor= IntVar()
var = IntVar()

def guardarInfo():
    
    global wb
    
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Datos de calibración"
    hoja1 = wb.create_sheet("Información de servicio")

    folioUsado=selectorFolio()
    materialUsado=selectorMaterial()
    gradoUsado=selectorGrado()
    usuarioUsado=selectorUsuario()
    revisorUsado=selectorRevisor()
    materialUsadoPatron=selectorMaterialPatron()

    hoja1.append(('Información general',''))
    hoja1.append(('Fecha de calibración:', time.strftime("%d/%m/%y")))
    hoja1.append(('Nº de Certificado LCM:', numeroCertificado.get()))
    hoja1.append(('Duración de la calibración:',''))
    hoja1.append(('Folios de Bitacora utilizados:', 'de '+folio1.get()+' a '+folio2.get()))
    hoja1.append(('Referencia de datos:', folioUsado+' '+folio1.get()+' a '+folio2.get()))
    hoja1.append(('Solicitante:', solicitante.get()))
    hoja1.append(('Número de certificado:', 'LCM '+numeroCertificado.get()))
    hoja1.append(('Número de solicitud:', numeroSolicitud.get()))
    hoja1.append(('Dirección del solicitante:', direccionSolicitante.get()))
    hoja1.append(('Lugar de la calibración:', 'Laboratorio de Metrología Dimensional - LCM'))
    hoja1.append(('', ''))

    hoja1.append(('Información del calibrando', ''))
    hoja1.append(('Objeto a calibrar:', 'Juego de bloques patrón'))
    hoja1.append(('Cantidad de bloques o instrumentos:', cantidadBloques.get()))
    hoja1.append(('Marca del instrumento:', marca.get()))
    hoja1.append(('N° de Serie del instrumento:', numeroSerie.get()))
    hoja1.append(('Material:', materialUsado))
    hoja1.append(('Modelo:', modelo.get()))
    hoja1.append(('Grado declarado:', gradoUsado))
    hoja1.append(('Identificación interna:', identInterna.get()))
    hoja1.append(('Coeficiente de expansión:', coefExp.get()))
    hoja1.append(('', ''))

    hoja1.append(('Responsable de la calibración:', usuarioUsado))
    hoja1.append(('Responsable de la revisión:', revisorUsado))
    hoja1.append(('', ''))

    hoja1.append(('Patrón a utilizar para la calibración:', patron.get()))
    hoja1.append(('Material de los bloques del Cliente:', materialUsado))
    hoja1.append(('Material de los bloques del Patrón:', materialUsadoPatron))

    ################## Guardar datos en documento para datos de búsqueda ##################

    wb2 = load_workbook(filename = 'DatosBusqueda.xlsx')
    work_sheet = wb2.active # Get active sheet
    work_sheet.append([numeroCertificado.get(), numeroSerie.get(), solicitante.get()])
        
    wb2.save('DatosBusqueda.xlsx')
    
    wb.save(numeroCertificado.get()+'.xlsx')

    FrameInfo.pack_forget()
    FrameCalibra.pack_forget()
    FrameBloques.pack(side="top", fill="both", expand=True)

################## Funciones de interaz ##################

def irCalibra():
    FrameInfo.pack_forget()
    FrameBloques.pack_forget()
    FrameCalibra.pack(side="top", fill="both", expand=True)

def irInfo():
    FrameBloques.pack_forget()
    FrameCalibra.pack_forget()
    FrameInfo.pack(side="top", fill="both", expand=True)

def irBloques():
    FrameInfo.pack_forget()
    FrameCalibra.pack_forget()
    FrameBloques.pack(side="top", fill="both", expand=True)
    
def obtenerSecuencia():
    Secuencia=secEscogida.get()
    Plantilla=planEscogida.get()
    return Secuencia,Plantilla

def crearHoja(Repeticiones):
    Secuencia,Plantilla=obtenerSecuencia()
    if Secuencia==1 and Plantilla==1: #Realiza secuencia completa1
        HojaDatosCom(Repeticiones)
    if Secuencia==1 and Plantilla==2: #Realiza secuencia completa2
        HojaDatosCom(Repeticiones)
    if Secuencia==2: #Realiza secuencia centros
        HojaDatosCen(Repeticiones)

# def enviarAlarma():
#     import smtplib, ssl
#     import yagmail
#     
#     receiver = "lrojas@lcm.go.cr"
#     body = "Hello there from Yagmail"
#     yag = yagmail.SMTP("alarma.calibracion@gmail.com")
#     yag.send(
#         to=receiver,
#         subject="Yagmail test with attachment",
#         contents=body, 
#         attachments=filename,
#     )

    port = 465 # For SSL
    smtp_server = "smtp.gmail.com"
    sender_email = "alarma_bloques@yahoo.com"  # Correo del que se envía alarma
    receiver_email = "lrojas@lcm.go.cr"  # Correo que recibe notificación
    password = "alarma.bloques"
    message = """\
    alarma.calibracion

    Proceso de calibracion termino exitosamente."""

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message)
   
def selectorSecuencia():

    global BloqueActual
    global dato
    global identificacionBloque
    global valorNominalBloque
    global wb
    global wb2
    
    listaMediciones=[]
    numerocertificado=numeroCertificado.get()
    cantidadbloques=cantidadBloques.get()
    tiempoinicial=tiempoInicial.get()
    tiempoestabilizacion=tiempoEstabilizacion.get()
    Repeticiones=repeticiones.get()
    
    wb = load_workbook(filename =numerocertificado+'.xlsx')
    hoja = wb.active
    
    Secuencia,Plantilla=obtenerSecuencia()
    
    DatoHumeVaisala=DatosVaisala() #Llama función Vaisala
    MedicionTemp1, MedicionTemp2, MedicionTemp3, MedicionTemp4=DatosFluke()
    
    if Secuencia==1 and Plantilla==1: #Realiza secuencia completa1
        listaMediciones, tiempoCorrida, t0=Completa1(tiempoinicial, tiempoestabilizacion, Repeticiones)
    if Secuencia==1 and Plantilla==2: #Realiza secuencia completa2
        listaMediciones, tiempoCorrida, t0=Completa2(tiempoinicial, tiempoestabilizacion, Repeticiones)
    if Secuencia==2: #Realiza secuencia centros
        listaMediciones, tiempoCorrida, t0=Centros(tiempoinicial, tiempoestabilizacion, Repeticiones)
        
#    enviarAlarma()
    
    MedicionTemp5, MedicionTemp6, MedicionTemp7, MedicionTemp8=DatosFluke()
    DatoHumeVaisala1=DatosVaisala() #Llama función Vaisala
    PromedioHumedad=(DatoHumeVaisala+DatoHumeVaisala1)/2

    listaDatos=[]

    listaDatos.append(BloqueActual) #Número de bloque
    listaDatos.append(time.strftime("%d/%m/%y")) #Fecha
    listaDatos.append(time.strftime("%H:%M:%S")) #Hora
    listaDatos.append(identificacionBloque[dato]) #Identificación del bloque
    listaDatos.append(valorNominalBloque[dato]) #Valor Nominal del bloque
    
    listaDatos.extend(listaMediciones)

    listaDatos.append(MedicionTemp1)
    listaDatos.append(MedicionTemp2)
    listaDatos.append(MedicionTemp3)
    listaDatos.append(MedicionTemp4)
    listaDatos.append(MedicionTemp5)
    listaDatos.append(MedicionTemp6)
    listaDatos.append(MedicionTemp7)
    listaDatos.append(MedicionTemp8)

    listaDatos.append(PromedioHumedad)

    if BloqueActual==1:
        crearHoja(Repeticiones)
        hoja.append(listaDatos)
        wb.save(numerocertificado+'.xlsx')
    elif BloqueActual==int(cantidadbloques):
        hoja.append(listaDatos)
        hoja1 = wb.active
        sheet_ranges = wb['Información de servicio']
        primerafecha=sheet_ranges['B2'].value
        primerafecha=primerafecha.split('/')
        dia1=primerafecha[0]
        mes1=primerafecha[1]    
        ultimafecha=time.strftime("%d/%m/%y")
        ultimafecha=ultimafecha.split('/')
        dia2=ultimafecha[0]
        mes2=ultimafecha[1]
        meses=int(mes2)-int(mes1)
        dias=int(dia2)-int(dia1)+30*meses
        #wb = load_workbook(filename =numerocertificado+'.xlsx')
        hoja1= wb['Información de servicio']
        hoja1["B4"]=dias+1
        wb.save(numerocertificado+'.xlsx')
    elif BloqueActual!=1 and BloqueActual!=int(cantidadbloques):
        hoja.append(listaDatos)
        wb.save(numerocertificado+'.xlsx')

    BloqueActual=BloqueActual+1
    dato=dato+1

def pausarCalibracion():
    global BloqueActual
    global dato
    global identificacionBloque
    global valorNominalBloque
    global wb4
    
    wb4 = openpyxl.Workbook()
    work_sheet = wb4.active
    work_sheet.title = "Hoja1"
    
    work_sheet.append(['Número de certificado', numeroCertificado.get()])
    work_sheet.append(['Cantidad de bloques', cantidadBloques.get()])
    #work_sheet.append(['Tiempo de estabilización', tiempoEstabilizacion.get()])
    work_sheet.append(['Repeticiones', repeticiones.get()])

    work_sheet.append(['Bloque actual', BloqueActual])
    work_sheet.append(['Dato', dato])

    work_sheet["D1"]='Identificación de los bloques'
    work_sheet["E1"]='Valor nominal de los bloques'

    
    for i in range(int(cantidadBloques.get())):
        work_sheet["D"+str(i+2)]=identificacionBloque[i]
        work_sheet["E"+str(i+2)]=valorNominalBloque[i]
        
    wb4.save('DatosPausa.xlsx')

def reanudarCalibracion():
    #gohome()
    global BloqueActual
    global dato
    global identificacionBloque
    global valorNominalBloque
    global wb5
    global wb

    wb5 = load_workbook(filename = 'DatosPausa.xlsx')
    sheet_ranges = wb5['Hoja1']
    
    listaMediciones=[]

    numerocertificado=sheet_ranges['B1'].value
    cantidadbloques=sheet_ranges['B2'].value
    tiempoestabilizacion=sheet_ranges['B3'].value
    tiempoinicial=tiempoInicial.get()
    Repeticiones=sheet_ranges['B4'].value

    wb = load_workbook(filename = numerocertificado+'.xlsx')
    hoja = wb.active
    
    Secuencia,Plantilla=obtenerSecuencia()
    DatoHumeVaisala=DatosVaisala() #Llama función Vaisala
    MedicionTemp1, MedicionTemp2, MedicionTemp3, MedicionTemp4=DatosFluke()
    
    if Secuencia==1 and Plantilla==1:                                          #Realiza secuencia completa1
        listaMediciones, tiempoCorrida, t0=Completa1(tiempoinicial, tiempoestabilizacion, Repeticiones)
    if Secuencia==1 and Plantilla==2:                                          #Realiza secuencia completa2
        listaMediciones, tiempoCorrida, t0=Completa2(tiempoinicial,tiempoestabilizacion, Repeticiones)
    if Secuencia==2:                                                           #Realiza secuencia centros
        listaMediciones, tiempoCorrida, t0=Centros(tiempoinicial,tiempoestabilizacion, Repeticiones)
    
#    enviarAlarma()
    
    MedicionTemp5, MedicionTemp6, MedicionTemp7, MedicionTemp8=DatosFluke()
    DatoHumeVaisala1=DatosVaisala()
                                                                               #Llama función Vaisala
    PromedioHumedad=(DatoHumeVaisala+DatoHumeVaisala1)/2

    listaDatos=[]

    listaDatos.append(BloqueActual)                                             #Número de bloque
    listaDatos.append(time.strftime("%d/%m/%y"))                                #Fecha
    listaDatos.append(time.strftime("%H:%M:%S"))                                #Hora
    listaDatos.append(identificacionBloque[dato])                              #Identificación del bloque
    listaDatos.append(valorNominalBloque[dato])                                #Valor Nominal del bloque

    tamaño=len(listaMediciones)

    for i in range (tamaño):
        listaDatos.append(listaMediciones[i])

    listaDatos.append(MedicionTemp1)
    listaDatos.append(MedicionTemp2)
    listaDatos.append(MedicionTemp3)
    listaDatos.append(MedicionTemp4)
    listaDatos.append(MedicionTemp5)
    listaDatos.append(MedicionTemp6)
    listaDatos.append(MedicionTemp7)
    listaDatos.append(MedicionTemp8)

    listaDatos.append(PromedioHumedad)

    if BloqueActual==1:
        crearHoja(Repeticiones)
        hoja.append(listaDatos)
    elif BloqueActual==int(cantidadbloques):
        hoja.append(listaDatos)
        wb.save(numerocertificado+'.xlsx')
    elif BloqueActual!=1 and BloqueActual!=int(cantidadbloques):
        hoja.append(listaDatos)

    BloqueActual=BloqueActual+1
    dato=dato+1
    
def recuperarDatos():
    global wb5
    
    wb5 = load_workbook(filename = 'DatosPausa.xlsx')
    sheet_ranges = wb5['Hoja1']

    global BloqueActual
    global dato
    global identificacionBloque
    global valorNominalBloque

    numerocertificado=sheet_ranges['B1'].value
    cantidadbloques=sheet_ranges['B2'].value
    tiempoestabilizacion=sheet_ranges['B3'].value
    Repeticiones=sheet_ranges['B4'].value
    BloqueActual=sheet_ranges['B5'].value
    dato=sheet_ranges['B6'].value
    
    fila1=2
    ident=''
    valorN=''
    
    while ident != None:
        ident=sheet_ranges['D'+str(fila1)].value
        if ident != None:
            identificacionBloque.append(ident)
        fila1=fila1+1

    fila2=2
    
    while valorN != None:
        valorN=sheet_ranges['E'+str(fila2)].value
        if valorN != None:
            valorNominalBloque.append(valorN)
        fila2=fila2+1
        
def selectorFolio():
    if escogerFolio.get()==1: #Folio1
        folioUsado='LRR-DI-01-FOLIOS'
    if escogerFolio.get()==2: #Folio2
        folioUsado='ORA-DI-01-FOLIOS'
    if escogerFolio.get()==3: #Folio3
        folioUsado='ILH-DI-03-FOLIOS'
    if escogerFolio.get()==4: #Otro
        folioUsado=nuevoFolio.get()
    return folioUsado

def selectorUnidades():
    if escogerUnidades.get()==1: #Unidad1
        unidadesUsadas='mm'
    if escogerUnidades.get()==2: #Unidad2
        unidadesUsadas='pulg'
    return unidadesUsadas

def selectorMaterial():
    if escogerMaterial.get()==1: #Material1
        materialUsado='Acero'
    if escogerMaterial.get()==2: #Material2
        materialUsado='Ceramicos'
    if escogerMaterial.get()==3: #Material3
        materialUsado='Carbono-Tungsteno'
    if escogerMaterial.get()==4: #Material4
        materialUsado='Carbono-Cromo'
    return materialUsado

def selectorMaterialPatron():
    if escogerMaterialPatron.get()==1: #Material1
        materialUsadoPatron='Acero'
    if escogerMaterialPatron.get()==2: #Material2
        materialUsadoPatron='Ceramicos'
    if escogerMaterialPatron.get()==3: #Material3
        materialUsadoPatron='Carbono-Tungsteno'
    if escogerMaterialPatron.get()==4: #Material4
        materialUsadoPatron='Carbono-Cromo'
    return materialUsadoPatron

def selectorGrado():
    if escogerGrado.get()==1: #Grado1
        gradoUsado='00 o K'
    if escogerGrado.get()==2: #Grado2
        gradoUsado='0'
    if escogerGrado.get()==3: #Grado3
        gradoUsado='1'
    if escogerGrado.get()==4: #Grado4
        gradoUsado='2'
    return gradoUsado

def selectorUsuario():
    if escogerUsuario.get()==1: #Usuario1
        usuarioUsado='Leonardo Rojas Rapso'
    if escogerUsuario.get()==2: #Usuario2
        usuarioUsado='Olman Ramos Alfaro'
    if escogerUsuario.get()==3: #Usuario3
        usuarioUsado='Ignacio López Hidalgo'
    if escogerUsuario.get()==4: #Usuario4
        usuarioUsado='Luis Chaves Santacruz'
    if escogerUsuario.get()==5: #Usuario5
        usuarioUsado='Francisco Sequeira'
    if escogerUsuario.get()==6: #Usuario6
        usuarioUsado='Luis Damian Rodriguez'
    if escogerUsuario.get()==7: #Usuario7
        usuarioUsado='Marcela Prendas Peña'
    if escogerUsuario.get()==8: #Usuario8
        usuarioUsado='Fernando Andrés Monge'
    if escogerUsuario.get()==9: #Otro
        usuarioUsado=nuevoUsuario.get()
    return usuarioUsado

def selectorRevisor():
    if escogerRevisor.get()==1: #Usuario1
        revisorUsado='Leonardo Rojas Rapso'
    if escogerRevisor.get()==2: #Usuario2
        revisorUsado='Olman Ramos Alfaro'
    if escogerRevisor.get()==3: #Usuario3
        revisorUsado='Ignacio López Hidalgo'
    if escogerRevisor.get()==4: #Usuario4
        revisorUsado='Luis Chaves Santacruz'
    if escogerRevisor.get()==5: #Usuario5
        revisorUsado='Francisco Sequeira'
    if escogerRevisor.get()==6: #Usuario6
        revisorUsado='Luis Damian Rodriguez'
    if escogerRevisor.get()==7: #Usuario7
        revisorUsado='Marcela Prendas Peña'
    if escogerRevisor.get()==8: #Usuario8
        revisorUsado='Fernando Andrés Monge'
    if escogerRevisor.get()==9: #Otro
        revisorUsado=nuevoRevisor.get()
    return revisorUsado


def agregarDato():
    global identificacionBloque
    global valorNominalBloque
    global count

    if count+1 <= int(cantidadBloques.get()):
        tabla.insert(parent='',index='end',iid = count,text='',values=(count+1,serie.get(),valorNominal.get()))
        count += 1
        identificacionBloque.append(serie.get())
        valorNominalBloque.append(valorNominal.get())
        serie.delete(0,END)
        valorNominal.delete(0,END)
    
    return identificacionBloque, valorNominalBloque

#Select Record
def seleccionarValores():
    global identificacionBloque
    global valorNominalBloque
    
    #clear entry boxes
    serie.delete(0,END)
    valorNominal.delete(0,END)
    
    #grab record
    selected=tabla.focus()
    #grab record values
    values = tabla.item(selected,'values')
    #temp_label.config(text=selected)

    #output to entry boxes
    serie.insert(0,values[1])
    valorNominal.insert(0,values[2])

#save Record
def editarValores():
    global identificacionBloque
    global valorNominalBloque
    
    selected=tabla.focus()

    values = tabla.item(selected,'values')
    
    #save new data 
    tabla.item(selected,text="",values=(values[0],serie.get(),valorNominal.get()))
    
    num=int(values[0])-1
    identificacionBloque[num]= serie.get()
    valorNominalBloque[num]= valorNominal.get()
    
   #clear entry boxes
    serie.delete(0,END)
    valorNominal.delete(0,END)

def imprimeBloqueActual():
    global count
    print(count) 
    return count

##def eliminarDato():
##    global identificacionBloque
##    global valorNominalBloque
##    global count
##    
##    tabla.insert(parent='',index='end',iid = count,text='',values=(count+1,serie.get(),valorNominal.get()))
##    count += 1
##    identificacionBloque.pop(count)
##    valorNominalBloque.pop(count)
##    return identificacionBloque, valorNominalBloque

def buscarDatos():
    
    global wb3
    
    wb3 = load_workbook(filename = 'DatosBusqueda.xlsx')
    sheet_ranges = wb3['Hoja1']
    fila=2
    cer=''
    ser=''
    emp=''

    while cer != None:
        cer=sheet_ranges['A'+str(fila)].value
        if cer == certificadoAnterior.get():
            val1=sheet_ranges['A'+str(fila)].value
            val2=sheet_ranges['B'+str(fila)].value
            val3=sheet_ranges['C'+str(fila)].value
            print('Número de certificado anterior:', val1, '  Número de serie del juego de bloques:', val2, '  Empresa:', val3)
        fila=fila+1
    
    while ser != None:
        ser=sheet_ranges['B'+str(fila)].value
        if ser == numeroSerie.get():
            val1=sheet_ranges['A'+str(fila)].value
            val2=sheet_ranges['B'+str(fila)].value
            val3=sheet_ranges['C'+str(fila)].value
            print('Número de certificado anterior:', val1, '  Número de serie del juego de bloques:', val2, '  Empresa:', val3)
        fila=fila+1

    while emp != None:
        emp=sheet_ranges['C'+str(fila)].value
        if emp == solicitante.get():
            val1=sheet_ranges['A'+str(fila)].value
            val2=sheet_ranges['B'+str(fila)].value
            val3=sheet_ranges['C'+str(fila)].value
            print('Número de certificado anterior:', val1, '  Número de serie del juego de bloques:', val2, '  Empresa:', val3)
        fila=fila+1

    certificadoAnterior.delete(0,END)

def cargarDatos():
    
    global wb1
    
    global identificacionBloque
    global valorNominalBloque
    
    wb1 = load_workbook(filename = archivoCargar.get()+'.xlsx')
    sheet_ranges = wb1['Datos de calibración']
    fila1=3
    ident=''
    valorN=''
    
    while ident != None:
        ident=sheet_ranges['D'+str(fila1)].value
        if ident != None:
            identificacionBloque.append(ident)
        fila1=fila1+1

    fila2=3
    
    while valorN != None:
        valorN=sheet_ranges['E'+str(fila2)].value
        if valorN != None:
            valorNominalBloque.append(valorN)
        fila2=fila2+1
    
    archivoCargar.delete(0,END)
    return identificacionBloque, valorNominalBloque

def limpiarDatos():
    global identificacionBloque
    global valorNominalBloque
    global count

    tabla.delete(*tabla.get_children())
    
    identificacionBloque=[]
    valorNominalBloque=[]
    count=0

def show_values():
    global esperaBloques
    esperaBloques=w1.get() #almacena para ser consultado el tiempo de espera
                           #de estabilización de los bloques
    print (w1.get())       # muestra el valor seleccionado en el slider de
                           # tiempo de estabilización del bloque
    
def crear_reporte():
    f= open("Indicadores.txt","w+")
    f.write("Laboratorio Costarricense de Metrología \n")
    f.write("Laboratorio de Dimensional \n")
    f.write("Indicadores de desempeño \n")
    today = date.today()
    f.write("Fecha de reporte: %s\r\n" % today)
    f.write("Número de certificado: %s\r\n" % numeroCertificado.get())

    f.write("Tiempo de calibración por bloque (min): %d\n" % 8.65)
    #minutos=tiempoCorrida/60
    #f.write("Tiempo de calibración por bloque: %d\r\n" % minutos)

    bloquespormes=int(cantidadBloques.get())*30/(int(cantidaddias.get()))
    f.write("Número de bloques calibrados por mes: %d\r\n" % bloquespormes)
    f.write("Cantidad de bloques fuera de clase: %s\r\n" % fueraClase.get())
    f.close()

################## Captura de datos TESA ##################

serTESA=serial.Serial('/dev/ttyUSBI', baudrate=1200, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN,
                          stopbits=serial.STOPBITS_TWO, xonxoff=True, timeout=0.5) #Configuración de puerto

def DatosTESA():                                   
    
    detenerse=0                     #Constante para while que captura dato
    def recv(serial):               #Definición de una función para recibir datos
        while True:
            
            data=serial.read(30)    #Lectura de 30 bytes
            if data == '':
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serTESA)          #Llamada de la función
        print(data)
        if data != b'':             #Comparación de datos recibidos, vacío hasta que se de la medición
            try:
                medicion=float(data) #Pasando de string a float
                MedicionBloque=medicion #Guardando dato en lista
            
            except:
                divisionDatos=data.split()
                print(divisionDatos)
                medicion=float(divisionDatos[1])    #Pasando de string a float #Cambio de valor a tomar por el TESA
                MedicionBloque=medicion #Guardando dato en lista
            detenerse = 1           #Condición para salir del while
    return MedicionBloque

################## Captura de datos Fluke ##################

serFluke=serial.Serial('/dev/ttyUSBK', baudrate=9600, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE,
                       stopbits=serial.STOPBITS_ONE, xonxoff=True, timeout=0.5) #Configuración de puerto


def DatosFluke():
    
    serFluke.write(b'READ?1\r\n') #Envío de instrucción para capturar dato de temperatura 1
    serFluke.write(b'READ?2\r\n') #Envío de instrucción para capturar dato de temperatura 2
    serFluke.write(b'READ?3\r\n') #Envío de instrucción para capturar dato de temperatura 3
    serFluke.write(b'READ?4\r\n') #Envío de instrucción para capturar dato de temperatura 4

    detenerse=0 #Constante para while que captura dato
    
    MedicionTemp1=0 #Creación de variable para almacenar mediciones de temperatura 1
    MedicionTemp2=0 #Creación de variable para almacenar mediciones de temperatura 2
    MedicionTemp3=0 #Creación de variable para almacenar mediciones de temperatura 3
    MedicionTemp4=0 #Creación de variable para almacenar mediciones de temperatura 4


    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(32) #Lectura de 32 bytes
            if data == '':
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serFluke) #Llamada de la función
        print(data)
        if data != b'': #Comparación de datos recibidos, vacío hasta que se de la medición
            todas=data.split()#Separar los 4 datos en una lista
            MedicionTemp1=float(todas[0]) #Guardando temperatura 1 en lista
            MedicionTemp2=float(todas[1]) #Guardando temperatura 2 en lista
            MedicionTemp3=float(todas[2]) #Guardando temperatura 3 en lista
            MedicionTemp4=float(todas[3]) #Guardando temperatura 4 en lista
            detenerse = 1  #Condición para salir del while
    return MedicionTemp1, MedicionTemp2, MedicionTemp3, MedicionTemp4

################## Captura de datos Vaisala ##################

serVaisala=serial.Serial('/dev/ttyUSBD', baudrate=4800, bytesize=serial.SEVENBITS,
                             parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5) #Configuración de puerto

def DatosVaisala():
    
    #serVaisala=serial.Serial('/dev/ttyAMA0', baudrate=4800, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout= 0.5)
    #serVaisala=serial.Serial('/dev/ttyUSB0', baudrate=19200, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout= 0.5)

    #serVaisala.write(b'RUN\r\n')
    #serVaisala.write(b'R\r/\n')
    #serVaisala.write(b'form "P=" 4.2 P " " U6 \t "T=" t " " U3 \t "RH=" 4.2 rh " " U5 \r\n')
    #serVaisala.write(b'form 4.2 rh " " \r\n')
    #serVaisala.write(b'form 4.2 P " " \t 4.2 t " " \t 4.2 rh " " \r\n')

    
    serVaisala.write(b'SEND\r\n') #Envío de instrucción para capturar datos del Vaisala
    
    detenerse=0 #Constante para while que captura dato
    
    #DatoPresVaisala=0 #Creación de variable para almacenar mediciones de presión atmosférica
    #DatoTempVaisala=0 #Creación de variable para almacenar mediciones de temperatura ambiente
    DatoHumeVaisala=0 #Creación de variable para almacenar mediciones de humedad relativa
    
    def recv(serial): #Definición de una función para recibir datos
        while True:
            data=serial.read(30)
            if data == '':
                continue
            else:
                break
            sleep(0.02)
        return data
    while detenerse == 0:
        data=recv(serVaisala)                   #Llamada de la función
        print(data)
        if data != b'':                         #Comparación de datos recibidos, vacío hasta que se de la medición
            todos=data.split()                  #Separar los 4 datos en una lista
            print(todos)
            #DatoPresVaisala=float(todos[1])     #Guardando presión atmosférica en lista
            #DatoTempVaisala=float(todos[2])     #Guardando temperatura en lista
            DatoHumeVaisala=float(todos[1])     #Guardando humedad relativa en lista
            detenerse = 1                       #Condición para salir del while
    return DatoHumeVaisala

    
FrameInfo=Frame(window)                         #Crea el frame donde estará la información de servicio
FrameBloques=Frame(window)                      #Crea el frame donde estarán los datos de los bloques
FrameCalibra=Frame(window)                      #Crea el frame donde se realizará el proceso de calibración


servo_pin = 26 #Pin que envía la señal al servomotor

def ActivaPedal(servo_pin): 

    myservotest = rpiservolib.SG90servo("servoone", 50, 2, 12) #Parámetros del servomotor

    myservotest.servo_move(servo_pin, 2.3, .5, False, .01)     #Movimiento a posición 2.3
    myservotest.servo_move(servo_pin, 7.5, .5, False, .01)     #Movimiento a posición 7.5

def Completa1(tiempoinicial, tiempoestabilizacion, Repeticiones):
    GPIO.output(EN_pin, GPIO.LOW)       #habilita los motores
    
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial
    
    listaMediciones=[]
    
    sleep(int(tiempoinicial)*60)
    
    for i in range(int(Repeticiones)):

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del patrón en posición 1 (centro)
        print(MedicionBloque)

        mymotortest1.motor_go(True, "Half", 417, .005, False, 2) #Mov de 1 a 2

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 2 (esquina)
        print(MedicionBloque)

        mymotortest1.motor_go(False, "Half", 96, .005, False, 2) #Mov1 de 2 a 3
        mymotortest2.motor_go(False, "Full", 398, .005, False, 1) #Mov2 de 2 a 3
        mymotortest1.motor_go(True, "Half", 178, .005, False, 1) #Mov3 de 2 a 3

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque) #Valor del calibrando en posición 3 (esquina)
        print(MedicionBloque)
            
        mymotortest1.motor_go(False, "Half", 178, .005, False, 2) #Mov de 3 a 4

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 4 (esquina)
        print(MedicionBloque)

        mymotortest2.motor_go(True, "Full", 796, .005, False, 2) #Mov de 4 a 5

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 5 (esquina)
        print(MedicionBloque)

        mymotortest1.motor_go(True, "Half", 174, .005, False, 2) #Mov de 5 a 6

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del calibrando en posición 6 (esquina)
        print(MedicionBloque)

        mymotortest1.motor_go(False, "Half", 174, .005, False, 2) #Mov de 6 a 5
        mymotortest2.motor_go(False, "Full", 398, .005, False, 1) #Mov de 5 a Esp2
        mymotortest1.motor_go(False, "Half", 330, .005, False, 1) #Mov de Esp2 a 1

        ActivaPedal(servo_pin) #Baja palpador
        sleep(int(tiempoestabilizacion)) #Tiempo de estabilización
        ActivaPedal(servo_pin) #Sube palpador
        MedicionBloque=DatosTESA() #Llama función TESA
        listaMediciones.append(MedicionBloque)  #Valor del patrón en posición 1 (centro)
        print(MedicionBloque)

    mymotortest1.motor_go(True, "Half", 208, .005, False, 2) #Mov de 1 a HOME

    ActivaPedal(servo_pin) #Baja palpador
    
    listaMediciones.append(MedicionBloque)
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                            #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    GPIO.output(EN_pin, GPIO.HIGH)       #Inhabilita los motores
    return listaMediciones, tiempoCorrida, t0
    

################## Secuencia completa - Plantilla2 ##################

def Completa2(tiempoinicial, tiempoestabilizacion, Repeticiones):
    GPIO.output(EN_pin, GPIO.LOW)       #habilita los motores
    
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial

    listaMediciones=[]
    
    sleep(int(tiempoinicial)*60)
    
    for i in range(int(Repeticiones)):

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)

        mymotortest1.motor_go(True, "Half", 416, .005, False, 2) #Mov de 1 a 2

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 2 (esquina)
        
        mymotortest1.motor_go(False, "Half", 96, .005, False, 2)    #Mov1 de 2 a 3
        mymotortest2.motor_go(False, "Full", 337, .005, False, 1)   #Mov2 de 2 a 3
        mymotortest1.motor_go(True, "Half", 182, .005, False, 1)    #Mov3 de 2 a 3

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 3 (esquina)
        
        mymotortest1.motor_go(False, "Half", 183, .005, False, 2)
                                                            #Mov de 3 a 4

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 4 (esquina)

        mymotortest2.motor_go(True, "Full", 683, .005, False, 2) #Mov de 4 a 5

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 5 (esquina)

        mymotortest1.motor_go(True, "Half", 178, .005, False, 2) #Mov de 5 a 6

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del calibrando en posición 6 (esquina)

        mymotortest1.motor_go(False, "Half", 178, .005, False, 2) #Mov de 6 a 5
        mymotortest2.motor_go(False, "Full", 342, .005, False, 1) #Mov de 5 a Esp2
        mymotortest1.motor_go(False, "Half", 332, .005, False, 1) #Mov de Esp2 a 1

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)

    mymotortest1.motor_go(True, "Half", 208, .005, False, 2) #Mov de 1 a HOME
        
    ActivaPedal(servo_pin)                              #Baja palpador
    listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)
    
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                                   #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    GPIO.output(EN_pin, GPIO.HIGH)       #Inhabilita los motores
    return listaMediciones, tiempoCorrida, t0

################## Secuencia centros ##################

def Centros(tiempoinicial, tiempoestabilizacion, Repeticiones):
    GPIO.output(EN_pin, GPIO.LOW)       #habilita los motores
    
    global valorNominalBloque
    global dato
    
    #obtenerAnguloBloque(valorNominalBloque[dato])          #Moverse a la siguiente pareja de bloques
    global t1
    t1=time.time()                                   #finaliza el conteo de espera de bloques
    tic=time.perf_counter()                                 #Toma el tiempo inicial
    
    listaMediciones=[]
    
    sleep(int(tiempoinicial)*60)
    
    for i in range(int(Repeticiones)):
        
                                                            #Se inicia en 1 y con palpador arriba

        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)              #Valor del patrón en posición 1 (centro)
     
        mymotortest1.motor_go(True, "Half", 407, .005, False, 2)
                                                            #Movimiento de punto1 a punto2
        
        ActivaPedal(servo_pin)                              #Baja palpador
        sleep(int(tiempoestabilizacion))                    #Tiempo de estabilización
        ActivaPedal(servo_pin)                              #Sube palpador
        MedicionBloque=DatosTESA()                   #Llama función TESA
        print(MedicionBloque)
        listaMediciones.append(MedicionBloque)               #Valor del calibrando en posición 2 (esquina)

        mymotortest1.motor_go(False, "Half", 410, .005, False, 2)

                                                            #Movimiento de punto2 a punto1
    
    ActivaPedal(servo_pin)                                  #Baja palpador
    sleep(int(tiempoestabilizacion))                        #Tiempo de estabilización
    ActivaPedal(servo_pin)                                  #Sube palpador
    MedicionBloque=DatosTESA()                       #Llama función TESA
    print(MedicionBloque)
    listaMediciones.append(MedicionBloque)                  #Valor del calibrando en posición 2 (esquina)

    mymotortest1.motor_go(True, "Half", 203, .005, False, 2)#Movimiento de punto1 a HOME

    ActivaPedal(servo_pin)                                  #Baja palpador
    #obtenerAnguloBloque(valorNominalBloques[dato])          #Moverse a la siguiente pareja de bloques
    toc=time.perf_counter()                                 #Toma el tiempo final
    global tiempoCorrida
    tiempoCorrida=toc-tic                                   #retorna el tiempo de corrida en segundos
    global t0
    t0=time.time()                                   #inicia el conteo de espera de bloques
    GPIO.output(EN_pin, GPIO.HIGH)       #Inhabilita los motores
    return listaMediciones, tiempoCorrida, t0

################## Frame Información de servicio ##################

FrameInfo.pack(side="top", fill="both", expand=True)        #Empieza el frame de información

Label(FrameInfo, text="Información servicio",font=(100)).place(x=10, y=10)

Label(FrameInfo, text="Información general",font=(100)).place(x=10, y=70)

Label(FrameInfo, text="Número de certificado: ").place(x=10, y=100)
numeroCertificado=Entry(FrameInfo)
numeroCertificado.place(x=180, y=100)

Label(FrameInfo, text="Solicitante: ").place(x=10, y=130)
solicitante=Entry(FrameInfo)
solicitante.place(x=180, y=130)

Label(FrameInfo, text="Número de solicitud: ").place(x=450, y=100)
numeroSolicitud=Entry(FrameInfo)
numeroSolicitud.place(x=620, y=100)

Label(FrameInfo, text="Dirección del solicitante: ").place(x=450, y=130)
direccionSolicitante=Entry(FrameInfo)
direccionSolicitante.place(x=620, y=130)

Label(FrameInfo, text="Folios de bitacora utilizados: de ").place(x=10, y=160)
folio1=Entry(FrameInfo, width=5)
folio1.place(x=230, y=160)
Label(FrameInfo, text="a ").place(x=290, y=160)
folio2=Entry(FrameInfo, width=5)
folio2.place(x=315, y=160)

Label(FrameInfo, text="Referencia de datos: ").place(x=10, y=190)
Radiobutton(FrameInfo, text="LRR-DI-01-FOLIOS", variable=escogerFolio,value=1, command=selectorFolio).place(x=10, y=210)
Radiobutton(FrameInfo, text="ORA-DI-01-FOLIOS", variable=escogerFolio,value=2, command=selectorFolio).place(x=10, y=230)
Radiobutton(FrameInfo, text="ILH-DI-03-FOLIOS", variable=escogerFolio,value=3, command=selectorFolio).place(x=10, y=250)
Radiobutton(FrameInfo, text="Otro:", variable=escogerFolio,value=4, command=selectorFolio).place(x=10, y=270)
nuevoFolio=Entry(FrameInfo)
nuevoFolio.place(x=180, y=270)

Label(FrameInfo, text="Información del calibrando",font=(100)).place(x=10, y=330)

Label(FrameInfo, text="Cantidad de bloques: ").place(x=10, y=360)
cantidadBloques=Entry(FrameInfo)
cantidadBloques.place(x=180, y=360)

Label(FrameInfo, text="Identificación interna: ").place(x=10, y=390)
identInterna=Entry(FrameInfo)
identInterna.place(x=180, y=390)

Label(FrameInfo, text="Número de serie: ").place(x=10, y=420)
numeroSerie=Entry(FrameInfo)
numeroSerie.place(x=180, y=420)

Label(FrameInfo, text="Marca: ").place(x=450, y=360)
marca=Entry(FrameInfo)
marca.place(x=620, y=360)

Label(FrameInfo, text="Modelo: ").place(x=450, y=390)
modelo=Entry(FrameInfo)
modelo.place(x=620, y=390)

Label(FrameInfo, text="Coeficiente de expansión: ").place(x=450, y=420)
coefExp=Entry(FrameInfo)
coefExp.place(x=620, y=420)

Label(FrameInfo, text="Unidades: ").place(x=10, y=450)
Radiobutton(FrameInfo, text="mm", variable=escogerUnidades,value=1, command=selectorUnidades).place(x=10, y=470)
Radiobutton(FrameInfo, text="pulg", variable=escogerUnidades,value=2, command=selectorUnidades).place(x=100, y=470)

Label(FrameInfo, text="Material: ").place(x=10, y=500)
Radiobutton(FrameInfo, text="Acero", variable=escogerMaterial,value=1, command=selectorMaterial).place(x=10, y=520)
Radiobutton(FrameInfo, text="Ceramicos", variable=escogerMaterial,value=2, command=selectorMaterial).place(x=10, y=540)
Radiobutton(FrameInfo, text="Carbono-Tungsteno", variable=escogerMaterial,value=3, command=selectorMaterial).place(x=10, y=560)
Radiobutton(FrameInfo, text="Carbono-Cromo", variable=escogerMaterial,value=4, command=selectorMaterial).place(x=10, y=580)

Label(FrameInfo, text="Grado declarado: ").place(x=450, y=500)
Radiobutton(FrameInfo, text="00 o K", variable=escogerGrado,value=1, command=selectorGrado).place(x=450, y=520)
Radiobutton(FrameInfo, text="0", variable=escogerGrado,value=2, command=selectorGrado).place(x=450, y=540)
Radiobutton(FrameInfo, text="1", variable=escogerGrado,value=3, command=selectorGrado).place(x=450, y=560)
Radiobutton(FrameInfo, text="2", variable=escogerGrado,value=4, command=selectorGrado).place(x=450, y=580)

Label(FrameInfo, text="Patrón a utilizar: ").place(x=890, y=100)
patron=Entry(FrameInfo)
patron.place(x=1060, y=100)

Label(FrameInfo, text="Material del patrón: ").place(x=890, y=130)
Radiobutton(FrameInfo, text="Acero", variable=escogerMaterialPatron,value=1, command=selectorMaterialPatron).place(x=890, y=150)
Radiobutton(FrameInfo, text="Ceramicos", variable=escogerMaterialPatron,value=2, command=selectorMaterialPatron).place(x=890, y=170)
Radiobutton(FrameInfo, text="Carbono-Tungsteno", variable=escogerMaterialPatron,value=3, command=selectorMaterialPatron).place(x=890, y=190)
Radiobutton(FrameInfo, text="Carbono-Cromo", variable=escogerMaterialPatron,value=4, command=selectorMaterialPatron).place(x=890, y=210)

Label(FrameInfo, text="Usuario: ").place(x=890, y=240)
Radiobutton(FrameInfo, text="Leonardo Rojas Rapso", variable=escogerUsuario,value=1, command=selectorUsuario).place(x=890, y=260)
Radiobutton(FrameInfo, text="Olman Ramos Alfaro", variable=escogerUsuario,value=2, command=selectorUsuario).place(x=890, y=280)
Radiobutton(FrameInfo, text="Ignacio López Hidalgo", variable=escogerUsuario,value=3, command=selectorUsuario).place(x=890, y=300)
Radiobutton(FrameInfo, text="Luis Chaves Santacruz", variable=escogerUsuario,value=4, command=selectorUsuario).place(x=890, y=320)
Radiobutton(FrameInfo, text="Francisco Sequeira", variable=escogerUsuario,value=5, command=selectorUsuario).place(x=890, y=340)
Radiobutton(FrameInfo, text="Luis Damian Rodriguez", variable=escogerUsuario,value=6, command=selectorUsuario).place(x=890, y=360)
Radiobutton(FrameInfo, text="Marcela Prendas Peña", variable=escogerUsuario,value=7, command=selectorUsuario).place(x=890, y=380)
Radiobutton(FrameInfo, text="Fernando Andrés Monge", variable=escogerUsuario,value=8, command=selectorUsuario).place(x=890, y=400)
Radiobutton(FrameInfo, text="Otro:", variable=escogerUsuario,value=9, command=selectorUsuario).place(x=890, y=420)
nuevoUsuario=Entry(FrameInfo)
nuevoUsuario.place(x=960, y=420)

Label(FrameInfo, text="Revisor: ").place(x=1200, y=240)
Radiobutton(FrameInfo, text="Leonardo Rojas Rapso", variable=escogerRevisor,value=1, command=selectorRevisor).place(x=1200, y=260)
Radiobutton(FrameInfo, text="Olman Ramos Alfaro", variable=escogerRevisor,value=2, command=selectorRevisor).place(x=1200, y=280)
Radiobutton(FrameInfo, text="Ignacio López Hidalgo", variable=escogerRevisor,value=3, command=selectorRevisor).place(x=1200, y=300)
Radiobutton(FrameInfo, text="Luis Chaves Santacruz", variable=escogerRevisor,value=4, command=selectorRevisor).place(x=1200, y=320)
Radiobutton(FrameInfo, text="Francisco Sequeira", variable=escogerRevisor,value=5, command=selectorRevisor).place(x=1200, y=340)
Radiobutton(FrameInfo, text="Luis Damian Rodriguez", variable=escogerRevisor,value=6, command=selectorRevisor).place(x=1200, y=360)
Radiobutton(FrameInfo, text="Marcela Prendas Peña", variable=escogerRevisor,value=7, command=selectorRevisor).place(x=1200, y=380)
Radiobutton(FrameInfo, text="Fernando Andrés Monge", variable=escogerRevisor,value=8, command=selectorRevisor).place(x=1200, y=400)
Radiobutton(FrameInfo, text="Otro:", variable=escogerRevisor,value=9, command=selectorRevisor).place(x=1200, y=420)
nuevoRevisor=Entry(FrameInfo)
nuevoRevisor.place(x=1270, y=420)

ttk.Button(FrameInfo, text='Crear hoja de datos', command=guardarInfo).place(x=1400, y=700) #Boton para crear hoja de datos

ttk.Button(FrameInfo, text='Reanudar calibración', command=irCalibra).place(x=1400, y=800)            #Boton para ir a frame de bloques

################## Frame Bloques ##################

tabla = ttk.Treeview(FrameBloques, height=20)
tabla.pack()

tabla['columns']= ('Número de Bloque','Número de Serie', 'Valor Nominal Bloque')
tabla.column("#0", width=0,  stretch=NO)
tabla.column("Número de Bloque",anchor=CENTER, width=150)
tabla.column("Número de Serie",anchor=CENTER, width=150)
tabla.column("Valor Nominal Bloque",anchor=CENTER, width=180)

tabla.heading("#0",text="",anchor=CENTER)
tabla.heading("Número de Bloque",text="Número de Bloque",anchor=CENTER)
tabla.heading("Número de Serie",text="Número de Serie",anchor=CENTER)
tabla.heading("Valor Nominal Bloque",text="Valor Nominal Bloque",anchor=CENTER)

serieLabel=Label(FrameBloques,text="Número de Serie")
serieLabel.place(x=610, y=450)

valorNominalLabel=Label(FrameBloques,text="Valor Nominal Bloque")
valorNominalLabel.place(x=800, y=450)

serie= Entry(FrameBloques)
serie.place(x=610, y=470)

valorNominal= Entry(FrameBloques)
valorNominal.place(x=800, y=470)

Label(FrameBloques, text="Número de certificado anterior: ").place(x=1100, y=30)
certificadoAnterior=Entry(FrameBloques)
certificadoAnterior.place(x=1320, y=30)

Label(FrameBloques, text="Nombre de archivo a cargar: ").place(x=1100, y=120)
Label(FrameBloques, text="(Sugerencia=Los nombres de archivo son el número de certificado)").place(x=1100, y=140)
archivoCargar=Entry(FrameBloques)
archivoCargar.place(x=1320, y=120)
Label(FrameBloques, text=".xlsx").place(x=1510, y=120)



global identificacionBloque
global valorNominalBloque
global count

identificacionBloque=[]
valorNominalBloque=[]
count=0

var.set(count)     # set it to 0 as the initial value

global BloqueActual
global dato
global avanceOrden
global tiempoCorrida
#global esperaBloques               #tiempo de estabilización de bloque cerca del comparador
#tiempoCorrida=DoubleVar()
#tiempoCorrida.set(0.00)
tiempoCorrida=0


#esperaBloques = DoubleVar()
#esperaBloques.set(0.00)
esperaBloques=0                     #Inicializamos el valor
BloqueActual=1
avanceOrden=0                       #inicializa el avance de la orden de trabajo
dato=0

#button
Input_button = Button(FrameBloques,text = "Agregar dato",command= agregarDato)
Input_button.pack()
Input_button.place(x=610, y=550)

select_button = Button(FrameBloques,text="Seleccionar valores a modificar", command=seleccionarValores)
select_button.pack()
select_button.place(x=750, y=550)

edit_button = Button(FrameBloques,text="Editar",command=editarValores)
edit_button.pack()
edit_button.place(x=830, y=600)

ttk.Button(FrameBloques, text='Buscar datos', command=buscarDatos).place(x=1100, y=60)       #Boton para crear hoja de datos

ttk.Button(FrameBloques, text='Cargar datos', command=cargarDatos).place(x=1100, y=170)      #Boton para crear hoja de datos

ttk.Button(FrameBloques, text='Limpiar datos', command=limpiarDatos).place(x=1100, y=230)    #Boton para crear hoja de datos

ttk.Button(FrameBloques, text='Siguiente', command=irCalibra).place(x=1400, y=800)          #Boton para ir a frame de calibración

ttk.Button(FrameBloques, text='Volver', command=irInfo).place(x=100, y=800)                  #Boton para ir a frame de información

################## Frame Calibración ##################

Label(FrameCalibra, text="Tiempo de estabilización inicial por bloque: ").place(x=10, y=10)
tiempoInicial=Entry(FrameCalibra)
tiempoInicial.place(x=300, y=10)
Label(FrameCalibra, text="min").place(x=487, y=10)

Label(FrameCalibra, text="Tiempo de estabilización para toma de cada medición: ").place(x=10, y=40)
tiempoEstabilizacion=Entry(FrameCalibra)
tiempoEstabilizacion.place(x=375, y=40)
Label(FrameCalibra, text="seg").place(x=562, y=40)

Label(FrameCalibra, text="Tiempo de estabilización de bloques: ").place(x=650, y=10)
w1 = Scale(FrameCalibra, from_=0, to=20,length= 450, tickinterval=1, orient=HORIZONTAL).place(x=650, y=30)
#w1.set(5)
ttk.Button(FrameCalibra, text='Fijar tiempo', command=show_values).place(x=1300, y=10)

Label(FrameCalibra, text="Repeticiones: ").place(x=10, y=70)
repeticiones=Entry(FrameCalibra)
repeticiones.place(x=105, y=70)

Label(FrameCalibra, text="Secuencia: ").place(x=10, y=130)
Radiobutton(FrameCalibra, text="Completa", variable=secEscogida,value=1, command=obtenerSecuencia).place(x=10, y=150)
Radiobutton(FrameCalibra, text="Centros", variable=secEscogida,value=2, command=obtenerSecuencia).place(x=10, y=170)

Label(FrameCalibra, text="Plantilla: ").place(x=10, y=210)
Radiobutton(FrameCalibra, text="1 (Bloque entre 10 mm y 100 mm)", variable=planEscogida,value=1, command=obtenerSecuencia).place(x=10, y=230)
Radiobutton(FrameCalibra, text="2 (Bloque menor o igual a 10 mm)", variable=planEscogida,value=2, command=obtenerSecuencia).place(x=10, y=250)

Label(FrameCalibra, text="Tiempo calibración por bloque (s): ").place(x=10, y=310)
Label(FrameCalibra, text=tiempoCorrida).place(x=250, y=310)

Label(FrameCalibra, text="Tiempo de espera de estabilización de bloques (s): ").place(x=10, y=380)
Label(FrameCalibra, text=esperaBloques).place(x=350, y=380)

Label(FrameCalibra, text="# Bloques fuera de clase: ").place(x=450, y=380)
fueraClase=Entry(FrameCalibra)
fueraClase.place(x=650, y=380)

Label(FrameCalibra, text="Avance de calibración (%): ").place(x=10, y=450)
Label(FrameCalibra, text=avanceOrden).place(x=250, y=450)

Label(FrameCalibra, text="Días de calibración: ").place(x=450, y=450)
cantidaddias=Entry(FrameCalibra)
cantidaddias.place(x=650, y=450)

ttk.Button(FrameCalibra, text='Generar Reporte', command=crear_reporte).place(x=930, y=450)
                                                                #Boton para ir a frame de calibración

Label(FrameCalibra, text="Puerto USB TESA: ").place(x=650, y=120)
Label(FrameCalibra, text="USBI").place(x=650, y=140)


Label(FrameCalibra, text="Puerto USB Fluke: ").place(x=800, y=120)
Label(FrameCalibra, text="USBK").place(x=800, y=140)

Label(FrameCalibra, text="Puerto USB Vaisala: ").place(x=950, y=120)
Label(FrameCalibra, text="USBD").place(x=950, y=140)

#Label(FrameCalibra, text=valorNominalBloque[dato]).place(x=930, y=400)

ttk.Button(FrameCalibra, text='Iniciar calibración', command=selectorSecuencia).place(x=100, y=600)
                                                                #Boton para ir a frame de calibración

ttk.Button(FrameCalibra, text='Pausar calibración', command=pausarCalibracion).place(x=530, y=600)
                                                                #Boton para parar secuencia

ttk.Button(FrameCalibra, text='Recuperar datos', command=recuperarDatos).place(x=930, y=600)
                                                                #Boton para parar secuencia

ttk.Button(FrameCalibra, text='Reanudar calibración', command=reanudarCalibracion).place(x=1300, y=600)
                                                                #Boton para reanudar secuencia

ttk.Button(FrameCalibra, text='Volver', command=irBloques).place(x=100, y=800)
                                                                #Boton para ir a frame de bloques

ttk.Button(window, text='Salir', command=quit).pack(side=BOTTOM)#Boton de salir

Button(FrameCalibra, text="Bloque actual", command=lambda: var.set(imprimeBloqueActual())).place(x=930, y=300)
Label(FrameCalibra, textvariable=var).place(x=930, y=330)

window.mainloop()

GPIO.cleanup()
